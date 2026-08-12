# core/views.py
from datetime import timedelta
from django.db import transaction
from django.db.models import F
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import viewsets, filters, status, generics
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone

from core.permissions import IsAdminOrReadOnlyForStaff
from core.services.dashoard.chart_service import get_dashboard_charts
from core.services.dashoard.credit_service import get_credit_summary
from core.services.dashoard.dashboard_followup import get_followups
from core.services.dashoard.dashboard_low_stock import get_low_stock
from core.services.dashoard.dashboard_order import get_orders
from core.services.dashoard.dashboard_staff_salary import get_staff_salaries
from core.services.uploads.order_upload import upload_order_excel
from core.services.uploads.purchase_upload import upload_purchase_excel
from core.services.uploads.sale_upload import upload_sales_excel
from core.services.uploads.stock_upload import upload_stock_excel
from core.services.uploads.new_mrp_upload import upload_mrp_excel

from .models import (
    Expense, SalaryTracker, SalaryTransaction, Supplier, Category, Stock,
    Purchase, PurchaseItem, Sale, SaleItem, FollowUpDashboard,
    Order, OrderItem, Staff, User, BikeSale, EmiTracker, BikeSaleFollowUp
)

from .serializers import (
    ExpenseSerializer, SalaryTrackerSerializer, SalaryTransactionSerializer,
    SaleReadSerializer, ServiceSaleReadSerializer, ServiceSaleSerializer,
    SupplierSerializer, CategorySerializer, StockSerializer, PurchaseSerializer,
    StockSaleSerializer, FollowUpDashboardSerializer, StaffSerializer,
    OrderSerializer, UserSerializer, BikeSaleSerializer,
    EmiTrackerSerializer, EmiTrackerUpdateSerializer, BikeSaleFollowUpSerializer
)

# =====================================================
# Base ViewSet to remove repetition
# =====================================================
class BaseModelViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    pagination_class = None


# ===================== CATEGORY =====================
class CategoryViewSet(BaseModelViewSet):
    queryset = Category.objects.all().order_by('name')
    serializer_class = CategorySerializer
    search_fields = ['name']
    ordering_fields = ['name']


# ===================== USER =====================
class UserViewSet(BaseModelViewSet):
    queryset = User.objects.all().order_by('id')
    serializer_class = UserSerializer


# ===================== SUPPLIER =====================
class SupplierViewSet(BaseModelViewSet):
    queryset = Supplier.objects.all().order_by('name')
    serializer_class = SupplierSerializer
    search_fields = ['name', 'contact', 'email']
    ordering_fields = ['name']


# ===================== STOCK =====================
class StockViewSet(BaseModelViewSet):
    queryset = Stock.objects.all().order_by('name')
    serializer_class = StockSerializer
    search_fields = ['name', 'model', 'category__name']
    ordering_fields = ['name', 'stock', 'purchase_price', 'sale_price']

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({"detail": "Stock item deleted successfully"}, status=status.HTTP_204_NO_CONTENT)


# ===================== PURCHASE =====================
class PurchaseViewSet(BaseModelViewSet):
    queryset = Purchase.objects.all().order_by('-date')
    serializer_class = PurchaseSerializer

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        purchase = self.get_object()
        purchase.delete()
        return Response({"detail": "Purchase deleted successfully"}, status=status.HTTP_204_NO_CONTENT)


# ===================== SALE =====================
class SaleViewSet(BaseModelViewSet):
    queryset = Sale.objects.all().order_by('-sale_date')

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            data = getattr(self.request, 'data', {})
            if data.get('is_servicing') or data.get('job_card_no'):
                return ServiceSaleSerializer
            return StockSaleSerializer
        if self.action in ['retrieve', 'list']:
            return SaleReadSerializer
        return StockSaleSerializer

    def list(self, request, *args, **kwargs):
        sales = self.get_queryset().select_related('handled_by')
        data = []
        for sale in sales:
            serializer = ServiceSaleReadSerializer(sale) if sale.is_servicing else SaleReadSerializer(sale)
            data.append(serializer.data)
        return Response(data)

    def retrieve(self, request, *args, **kwargs):
        sale = self.get_object()
        serializer = ServiceSaleReadSerializer(sale) if sale.is_servicing else SaleReadSerializer(sale)
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        sale = serializer.save()
        read_serializer = ServiceSaleReadSerializer(sale) if sale.is_servicing else SaleReadSerializer(sale)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        sale = serializer.save()
        read_serializer = ServiceSaleReadSerializer(sale) if sale.is_servicing else SaleReadSerializer(sale)
        return Response(read_serializer.data)


# =====================================================
# Dashboard APIs
# =====================================================
@api_view(["GET"])
def dashboard_charts_api(request):
    period = request.GET.get("period", "monthly")
    return Response(get_dashboard_charts(period))


@api_view(["GET"])
def dashboard_credit_api(request):
    period = request.GET.get("period", "monthly")
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 5))
    return Response(get_credit_summary(period=period, page=page, page_size=page_size))


@api_view(["GET"])
def followups_api(request):
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 5))
    days = int(request.GET.get("days", 10))
    return Response(get_followups(days=days, page=page, page_size=page_size))


@api_view(["GET"])
def low_stock_api(request):
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 5))
    threshold = int(request.GET.get("threshold", 5))
    return Response(get_low_stock(threshold=threshold, page=page, page_size=page_size))


@api_view(["GET"])
def orders_api(request):
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 5))
    return Response(get_orders(page=page, page_size=page_size))


@api_view(["GET"])
def staff_salary_api(request):
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 5))
    return Response(get_staff_salaries(page=page, page_size=page_size))


# ===================== STAFF =====================
class StaffViewSet(BaseModelViewSet):
    queryset = Staff.objects.all().order_by('name')
    serializer_class = StaffSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnlyForStaff]


# ===================== FOLLOW-UP DASHBOARD =====================
class FollowUpDashboardViewSet(viewsets.ViewSet):
    """
    Unified follow-up dashboard: Sale + BikeSale follow-ups
    """

    def list(self, request):
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))

        sale_qs = FollowUpDashboard.objects.values(
            'id', 'customer_name', 'contact_no', 'vehicle', 'delivery_date',
            'post_service_feedback_date', 'follow_up_date', 'remarks', 'status'
        )
        bike_qs = BikeSaleFollowUp.objects.values(
            'id', 'customer_name', 'contact_no', 'vehicle', 'delivery_date',
            'post_service_feedback_date', 'follow_up_date', 'remarks', 'status'
        )

        combined = list(sale_qs) + list(bike_qs)
        combined.sort(key=lambda x: x['follow_up_date'] or x['delivery_date'])

        paginator = Paginator(combined, page_size)
        page_obj = paginator.get_page(page)
        return Response({
            "results": list(page_obj),
            "pagination": {
                "page": page_obj.number,
                "page_size": page_size,
                "total_pages": paginator.num_pages,
                "total_items": paginator.count,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        })

    @action(detail=True, methods=['post'], url_path='terminate')
    def terminate_followup(self, request, pk=None):
        followup = (
            FollowUpDashboard.objects.filter(pk=pk).first() or
            BikeSaleFollowUp.objects.filter(pk=pk).first()
        )
        if not followup:
            return Response({"detail": "Follow-up not found."}, status=status.HTTP_404_NOT_FOUND)

        reason = request.data.get('reason')
        followup.terminate(reason=reason)
        serializer = FollowUpDashboardSerializer(followup)
        return Response(serializer.data)


# ===================== ORDER =====================
class OrderViewSet(BaseModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer


# ===================== SALARY =====================
class SalaryTrackerViewSet(BaseModelViewSet):
    queryset = SalaryTracker.objects.all()
    serializer_class = SalaryTrackerSerializer


class SalaryTransactionViewSet(BaseModelViewSet):
    queryset = SalaryTransaction.objects.all()
    serializer_class = SalaryTransactionSerializer


# ===================== EXPENSE =====================
class ExpenseViewSet(BaseModelViewSet):
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer


# ===================== EXCEL UPLOAD APIS =====================
def handle_excel_upload(file, upload_func, user=None):
    if not file:
        return Response({"error": "Excel file required"}, status=400)
    return Response(upload_func(file, user) if user else upload_func(file))


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def purchase_excel_upload_api(request):
    return handle_excel_upload(request.FILES.get("file"), upload_purchase_excel, request.user)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def new_mrp_excel_upload_api(request):
    return handle_excel_upload(request.FILES.get("file"), upload_mrp_excel)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sale_excel_upload_api(request):
    return handle_excel_upload(request.FILES.get("file"), upload_sales_excel)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def stock_excel_upload_api(request):
    return handle_excel_upload(request.FILES.get("file"), upload_stock_excel)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def order_excel_upload_api(request):
    return handle_excel_upload(request.FILES.get("file"), upload_order_excel)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stock_excel_export_api(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = ["Item No", "Name", "Model", "Category", "Quantity", "MRP"]

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # tenant/company filter छ भने थप्नुहोस्, जस्तै:
    # stocks = Stock.objects.filter(company=request.user.company).select_related('category')
    stocks = Stock.objects.all().select_related('category')

    for row_num, stock in enumerate(stocks, start=2):
        ws.cell(row=row_num, column=1, value=stock.item_no)
        ws.cell(row=row_num, column=2, value=stock.name)
        ws.cell(row=row_num, column=3, value=getattr(stock, 'model', '') or '')
        ws.cell(row=row_num, column=4, value=stock.category.name if stock.category else '')
        ws.cell(row=row_num, column=5, value=stock.stock)
        ws.cell(row=row_num, column=6, value=stock.purchase_price)
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(len(header) + 4, 15)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="stock_export.xlsx"'
    wb.save(response)
    return response


# ===================== BIKE SALE =====================
class BikeSaleViewSet(BaseModelViewSet):
    queryset = BikeSale.objects.all().order_by('-sale_date')
    serializer_class = BikeSaleSerializer


# ===================== EMI TRACKER =====================
class EmiTrackerViewSet(BaseModelViewSet):
    queryset = EmiTracker.objects.all().order_by('installment_no')
    serializer_class = EmiTrackerSerializer

    def create(self, request, *args, **kwargs):
        sale_id = request.data.get('sale')
        sale = BikeSale.objects.get(id=sale_id)
        if sale.sale_type != 'emi':
            return Response({"error": "EMI tracker can only be created for EMI sales."}, status=status.HTTP_400_BAD_REQUEST)
        return super().create(request, *args, **kwargs)


class EmiTrackerUpdateAPIView(generics.UpdateAPIView):
    queryset = EmiTracker.objects.all()
    serializer_class = EmiTrackerUpdateSerializer
    lookup_field = 'id'  # frontend sends EMI id